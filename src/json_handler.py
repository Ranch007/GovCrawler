"""
JSON 操作 — 读写、去重合并、正文清理、CSV/XLSX 导出。

每篇政策为一个 JSON 对象:
    {"pcode", "title", "pubtimeStr", "url", "childtype", "puborg", "post"}
"""

import csv
import json
import os
import re
from typing import List, Dict

# 正文可能很长，提高 CSV 字段上限（默认 128KB → 10MB）
csv.field_size_limit(10 * 1024 * 1024)


MATCH_KW_KEY = "匹配关键词"
# 导出列顺序（CSV / XLSX 共用）
EXPORT_FIELDS = ["序号", "pcode", "title", "pubtimeStr", "url", "childtype", "puborg", "post", MATCH_KW_KEY]


def write_json(filepath: str, items: List[Dict[str, str]]) -> None:
    """
    写入 JSON 文件（UTF-8，indent=2）。

    Args:
        filepath: 输出路径
        items:    政策对象列表 [{pcode, title, ...}, ...]
    """
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(items, f, ensure_ascii=False, indent=2)
    print(f"  [JSON] 写入 {len(items)} 篇政策 -> {filepath}")


def read_json(filepath: str) -> List[Dict[str, str]]:
    """
    读取 JSON 文件，返回政策对象列表。

    Returns:
        [{pcode, title, ...}, ...]
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"文件不存在: {filepath}")

    with open(filepath, "r", encoding="utf-8") as f:
        data = json.load(f)
    return data if isinstance(data, list) else []


def merge_json(json_paths: List[str], output_path: str) -> str:
    """
    合并多个关键词 JSON，按 pcode + title 双重去重，聚合"匹配关键词"。

    去重规则:
      - pcode + title（strip后）相同 → 去重，保留首次出现的条目
      - 聚合"匹配关键词"字段，用逗号分隔命中该政策的所有关键词
      - post 字段置为空字符串，等待 Phase 3 爬取填充

    Args:
        json_paths:  各关键词的 JSON 文件路径列表
        output_path: 合并后输出路径

    Returns:
        输出文件路径
    """
    seen_keys: Dict[str, int] = {}
    merged_items: List[Dict[str, str]] = []

    for filepath in json_paths:
        if not os.path.exists(filepath):
            print(f"  [WARN] 跳过不存在的文件: {filepath}")
            continue

        kw_name = os.path.splitext(os.path.basename(filepath))[0]

        with open(filepath, "r", encoding="utf-8") as f:
            items = json.load(f)
        if not isinstance(items, list):
            continue

        for item in items:
            pcode = item.get("pcode", "").strip()
            title = item.get("title", "").strip()
            if not title:
                continue
            item_key = f"{pcode}|{title}"
            if item_key in seen_keys:
                idx = seen_keys[item_key]
                existing = merged_items[idx].get(MATCH_KW_KEY, "")
                kw_set = set(k.strip() for k in existing.split(",") if k.strip())
                kw_set.add(kw_name)
                merged_items[idx][MATCH_KW_KEY] = ", ".join(sorted(kw_set))
            else:
                seen_keys[item_key] = len(merged_items)
                item[MATCH_KW_KEY] = kw_name
                item.setdefault("post", "")
                merged_items.append(item)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(merged_items, f, ensure_ascii=False, indent=2)

    print(f"  [JSON] 合并去重: {len(json_paths)} 个文件 -> {len(merged_items)} 篇（pcode+title去重） -> {output_path}")
    return output_path


_RE_BLANK = re.compile(r"\s+")
_FOOTER_MARKER = "扫一扫在手机打开当前页"


def strip_footer(text: str) -> str:
    """切除"扫一扫在手机打开当前页"及其后所有内容"""
    idx = text.find(_FOOTER_MARKER)
    if idx != -1:
        text = text[:idx]
    return text.strip()


def sanitize_post(text: str) -> str:
    """清理 post：切除页脚，移除所有空白字符（空格、换行、制表符等）。

    在写入 CSV / XLSX 之前调用，保证不含换行和冗余页脚。
    """
    if not text:
        return ""
    text = strip_footer(text)
    return _RE_BLANK.sub("", text)


def json_to_csv(input_json: str, output_csv: str) -> str:
    """
    将合并后的 JSON 转换为 CSV（UTF-8 with BOM）。

    CSV 列: 序号 + API 原始字段 + post + 匹配关键词
    序号从 1 开始递增。
    post 字段中的换行符会被替换为空格，确保每条政策占一行。

    Args:
        input_json:  合并后的 JSON 路径（如 allKeyword.json）
        output_csv:  输出 CSV 路径

    Returns:
        输出 CSV 路径
    """
    items = read_json(input_json)

    with open(output_csv, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=EXPORT_FIELDS, extrasaction="ignore")
        writer.writeheader()
        for i, item in enumerate(items, 1):
            row = {**item, "序号": str(i)}
            row["post"] = sanitize_post(row.get("post", ""))
            writer.writerow(row)

    has_post = sum(1 for it in items if it.get("post", "").strip())
    print(f"  [CSV] 导出完成: {len(items)} 行（含正文 {has_post} 篇） -> {output_csv}")
    return output_csv


def json_to_xlsx(input_json: str, output_xlsx: str) -> str:
    """
    将合并后的 JSON 转换为 XLSX（openpyxl，表头加粗冻结首行）。

    Args:
        input_json:  合并后的 JSON 路径（如 allKeyword.json）
        output_xlsx: 输出 XLSX 路径

    Returns:
        输出 XLSX 路径
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    items = read_json(input_json)

    wb = Workbook()
    ws = wb.active
    ws.title = "政策数据"

    # 表头
    header_font = Font(bold=True)
    for col_idx, field in enumerate(EXPORT_FIELDS, 1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.font = header_font

    # 数据行
    for i, item in enumerate(items, 1):
        row = {**item, "序号": str(i)}
        post = sanitize_post(row.get("post", ""))
        for col_idx, field in enumerate(EXPORT_FIELDS, 1):
            val = post if field == "post" else row.get(field, "")
            ws.cell(row=i + 1, column=col_idx, value=val)

    # 调整列宽（post 列给较大宽度）
    col_widths = {
        "序号": 6, "pcode": 22, "title": 50, "pubtimeStr": 12,
        "url": 40, "childtype": 14, "puborg": 16, "post": 80,
        MATCH_KW_KEY: 14,
    }
    for col_idx, field in enumerate(EXPORT_FIELDS, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = col_widths.get(field, 14)

    # 冻结首行
    ws.freeze_panes = "A2"

    wb.save(output_xlsx)
    print(f"  [XLSX] 导出完成: {len(items)} 行 -> {output_xlsx}")
    return output_xlsx